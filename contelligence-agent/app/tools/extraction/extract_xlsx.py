"""Tool for extracting sheet data from Excel XLSX files."""

from __future__ import annotations

import base64
import io
import logging
import pathlib
from typing import Any, Literal

from pydantic import BaseModel, Field

from app.core.tool_registry import define_tool

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Parameters
# ---------------------------------------------------------------------------

class ExtractXlsxParams(BaseModel):
    """Parameters for the extract_xlsx tool."""

    file_bytes: bytes | None = Field(
        None,
        description="Raw bytes of the XLSX file. Preferred when the caller already has the content in memory.",
    )
    file_bytes_b64: str | None = Field(
        None,
        description="Base64-encoded XLSX file content.",
    )
    local_path: str | None = Field(
        None,
        description="Absolute or relative path to an XLSX file on the local filesystem.",
    )
    storage_account: str | None = Field(
        default=None,
        description=(
            "Azure Storage account name to connect to. "
            "Provide this to read from a storage account "
            "(authentication uses DefaultAzureCredential)."
        ),
    )
    container: str | None = Field(
        None, description="Azure Blob Storage container name."
    )
    path: str | None = Field(
        None, description="Blob path to the XLSX file."
    )
    sheets: str | None = Field(
        None,
        description=(
            "Comma-separated sheet names to extract. "
            "Omit to extract all sheets."
        ),
    )
    format: Literal["markdown", "json"] = Field(
        "markdown",
        description="Output format: 'markdown' returns a rendered markdown document, 'json' returns structured data.",
    )
    filename: str | None = Field(
        None,
        description="Optional descriptive filename for the result metadata.",
    )


# ---------------------------------------------------------------------------
# Tool definition
# ---------------------------------------------------------------------------

@define_tool(
    name="extract_xlsx",
    description=(
        "Extract tabular data from an Excel XLSX workbook. Accepts the file "
        "as raw bytes (file_bytes), base64-encoded bytes (file_bytes_b64), a "
        "local filesystem path (local_path), or reads from Azure Blob Storage "
        "(storage_account + container + path). Returns a markdown document (default) or "
        "structured JSON. Uses openpyxl to read sheets, headers, and rows. "
        "Supports filtering by sheet name."
    ),
    parameters_model=ExtractXlsxParams,
)
async def extract_xlsx(params: ExtractXlsxParams, context: dict) -> dict[str, Any]:
    """Extract sheet data from an XLSX provided as bytes, local path, or fetched from blob storage."""
    try:
        from openpyxl import load_workbook

        # Resolve the file bytes from the supplied source.
        if params.file_bytes:
            logger.info("Parsing XLSX from supplied raw bytes")
            xlsx_bytes: bytes = params.file_bytes
        elif params.file_bytes_b64:
            logger.info("Parsing XLSX from base64-encoded bytes")
            xlsx_bytes = base64.b64decode(params.file_bytes_b64)
        elif params.local_path:
            resolved = pathlib.Path(params.local_path).expanduser().resolve()
            if not resolved.is_file():
                return {
                    "error": f"Local file not found: {resolved}",
                    "filename": params.filename or params.local_path,
                }
            logger.info("Reading XLSX from local path %s", resolved)
            xlsx_bytes = resolved.read_bytes()
        elif params.storage_account and params.container and params.path:
            from app.connectors.blob_connector import BlobConnectorAdapter
            try:
                connector = BlobConnectorAdapter(
                    account_name=params.storage_account,
                    credential_type="default_azure_credential",
                )
                logger.info("Downloading XLSX blob %s/%s", params.container, params.path)
                xlsx_bytes = await connector.download_blob(params.container, params.path)
            except Exception as blob_err:
                return {
                    "error": f"Failed to download XLSX from blob storage: {blob_err}",
                    "filename": params.filename or params.path,
                }
            finally:
                if connector is not None:
                    await connector.close()
        else:
            return {
                "error": (
                    "Provide one of: file_bytes, file_bytes_b64, local_path, "
                    "or storage_account + container + path."
                ),
                "filename": params.filename or "",
            }

        wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)

        # Determine which sheets to process.
        if params.sheets:
            requested = [s.strip() for s in params.sheets.split(",")]
            sheet_names = [s for s in requested if s in wb.sheetnames]
        else:
            sheet_names = wb.sheetnames

        sheets_data: list[dict[str, Any]] = []
        for name in sheet_names:
            ws = wb[name]
            rows: list[list[Any]] = []
            for row in ws.iter_rows(values_only=True):
                rows.append([cell for cell in row])

            if rows:
                headers = [str(h) if h is not None else "" for h in rows[0]]
                data_rows = [
                    [str(c) if c is not None else "" for c in row]
                    for row in rows[1:]
                ]
            else:
                headers = []
                data_rows = []

            sheets_data.append(
                {
                    "name": name,
                    "headers": headers,
                    "rows": data_rows,
                    "row_count": len(data_rows),
                }
            )

        wb.close()

        label = params.filename or params.local_path or params.path or "unknown.xlsx"

        if params.format == "markdown":
            return {
                "filename": label,
                "content": _xlsx_to_markdown(label, sheets_data),
            }

        return {
            "filename": label,
            "sheets": sheets_data,
        }

    except Exception as exc:
        label = params.filename or params.local_path or params.path or "unknown.xlsx"
        logger.exception("extract_xlsx failed for %s", label)
        return {"error": str(exc), "filename": label}


# ---------------------------------------------------------------------------
# Markdown renderer
# ---------------------------------------------------------------------------

def _xlsx_to_markdown(label: str, sheets: list[dict[str, Any]]) -> str:
    """Convert extracted XLSX sheet data into a readable markdown document."""
    parts: list[str] = [f"# {label}", ""]

    for sheet in sheets:
        name = sheet["name"]
        headers = sheet.get("headers", [])
        rows = sheet.get("rows", [])
        row_count = sheet.get("row_count", len(rows))

        parts.append(f"## {name}")
        parts.append("")
        parts.append(f"*{row_count} data rows*")
        parts.append("")

        if headers:
            parts.append("| " + " | ".join(headers) + " |")
            parts.append("| " + " | ".join("---" for _ in headers) + " |")
            for row in rows:
                parts.append("| " + " | ".join(row) + " |")
            parts.append("")

    return "\n".join(parts).rstrip() + "\n"
